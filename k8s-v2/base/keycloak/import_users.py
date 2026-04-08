#!/usr/bin/env python3
"""Import users into Keycloak from CSV or XLSX.

Expected columns:
- Nom
- Prenom
- Email
- Password

For each user, the script:
- creates or updates the user
- enables the account
- sets the initial password
- forces UPDATE_PASSWORD at first login
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


REQUIRED_COLUMNS = {"Nom", "Prenom", "Email", "Password"}


class KeycloakError(RuntimeError):
    pass


@dataclass
class UserRow:
    last_name: str
    first_name: str
    email: str
    password: str

    @property
    def username(self) -> str:
        # In this project we use the email as the login identifier.
        return self.email


class KeycloakAdminClient:
    def __init__(self, base_url: str, realm: str, admin_user: str, admin_password: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.realm = realm
        self.admin_user = admin_user
        self.admin_password = admin_password
        self._token = self._get_admin_token()

    def _request(
        self,
        method: str,
        path: str,
        *,
        data: dict | list | None = None,
        form: dict[str, str] | None = None,
        expected: set[int] | None = None,
        extra_headers: dict[str, str] | None = None,
    ) -> tuple[int, str, dict[str, str]]:
        url = f"{self.base_url}{path}"
        headers = extra_headers.copy() if extra_headers else {}
        body = None

        if data is not None:
            body = json.dumps(data).encode("utf-8")
            headers["Content-Type"] = "application/json"
        elif form is not None:
            body = urllib.parse.urlencode(form).encode("utf-8")
            headers["Content-Type"] = "application/x-www-form-urlencoded"

        request = urllib.request.Request(url, data=body, method=method, headers=headers)

        try:
            with urllib.request.urlopen(request) as response:
                payload = response.read().decode("utf-8")
                status = response.getcode()
                response_headers = dict(response.headers.items())
        except urllib.error.HTTPError as exc:
            payload = exc.read().decode("utf-8", errors="replace")
            response_headers = dict(exc.headers.items())
            status = exc.code
            if expected is None or status not in expected:
                raise KeycloakError(f"{method} {url} failed with {status}: {payload}") from exc
        except urllib.error.URLError as exc:
            raise KeycloakError(f"Cannot reach Keycloak at {url}: {exc}") from exc

        if expected is not None and status not in expected:
            raise KeycloakError(f"{method} {url} returned {status}: {payload}")

        return status, payload, response_headers

    def _get_admin_token(self) -> str:
        _, payload, _ = self._request(
            "POST",
            f"/realms/master/protocol/openid-connect/token",
            form={
                "client_id": "admin-cli",
                "grant_type": "password",
                "username": self.admin_user,
                "password": self.admin_password,
            },
            expected={200},
        )
        token = json.loads(payload)["access_token"]
        return token

    @property
    def auth_headers(self) -> dict[str, str]:
        return {"Authorization": f"Bearer {self._token}"}

    def find_user_by_email(self, email: str) -> dict | None:
        _, payload, _ = self._request(
            "GET",
            f"/admin/realms/{urllib.parse.quote(self.realm)}/users?email={urllib.parse.quote(email)}&exact=true",
            expected={200},
            extra_headers=self.auth_headers,
        )
        users = json.loads(payload)
        return users[0] if users else None

    def create_user(self, user: UserRow) -> str:
        _, _, headers = self._request(
            "POST",
            f"/admin/realms/{urllib.parse.quote(self.realm)}/users",
            data={
                "username": user.username,
                "email": user.email,
                "firstName": user.first_name,
                "lastName": user.last_name,
                "enabled": True,
                "emailVerified": True,
                "requiredActions": ["UPDATE_PASSWORD"],
            },
            expected={201},
            extra_headers=self.auth_headers,
        )
        location = headers.get("Location", "")
        user_id = location.rstrip("/").split("/")[-1]
        if not user_id:
            raise KeycloakError(f"Could not determine created user id for {user.email}")
        return user_id

    def update_user(self, user_id: str, user: UserRow) -> None:
        self._request(
            "PUT",
            f"/admin/realms/{urllib.parse.quote(self.realm)}/users/{urllib.parse.quote(user_id)}",
            data={
                "username": user.username,
                "email": user.email,
                "firstName": user.first_name,
                "lastName": user.last_name,
                "enabled": True,
                "emailVerified": True,
                "requiredActions": ["UPDATE_PASSWORD"],
            },
            expected={204},
            extra_headers=self.auth_headers,
        )

    def set_password(self, user_id: str, password: str) -> None:
        self._request(
            "PUT",
            f"/admin/realms/{urllib.parse.quote(self.realm)}/users/{urllib.parse.quote(user_id)}/reset-password",
            data={
                "type": "password",
                "value": password,
                "temporary": True,
            },
            expected={204},
            extra_headers=self.auth_headers,
        )


def load_csv(path: Path) -> list[UserRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise KeycloakError(f"{path} has no header row")
        missing = REQUIRED_COLUMNS.difference(reader.fieldnames)
        if missing:
            raise KeycloakError(f"{path} is missing required columns: {', '.join(sorted(missing))}")

        rows = []
        for index, raw_row in enumerate(reader, start=2):
            rows.append(parse_row(raw_row, path, index))
        return rows


def load_xlsx(path: Path) -> list[UserRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise KeycloakError(
            "XLSX import requires openpyxl. Install it in a virtualenv or use the CSV template instead."
        ) from exc

    workbook = load_workbook(filename=path, read_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        raise KeycloakError(f"{path} is empty")

    header = [str(value).strip() if value is not None else "" for value in values[0]]
    missing = REQUIRED_COLUMNS.difference(header)
    if missing:
        raise KeycloakError(f"{path} is missing required columns: {', '.join(sorted(missing))}")

    rows = []
    for index, row_values in enumerate(values[1:], start=2):
        row = {header[pos]: ("" if value is None else str(value)) for pos, value in enumerate(row_values)}
        rows.append(parse_row(row, path, index))
    return rows


def parse_row(raw_row: dict[str, str], path: Path, index: int) -> UserRow:
    last_name = raw_row["Nom"].strip()
    first_name = raw_row["Prenom"].strip()
    email = raw_row["Email"].strip().lower()
    password = raw_row["Password"].strip()

    if not all([last_name, first_name, email, password]):
        raise KeycloakError(f"{path}:{index} has empty required values")

    if "@" not in email:
        raise KeycloakError(f"{path}:{index} has an invalid email: {email}")

    return UserRow(
        last_name=last_name,
        first_name=first_name,
        email=email,
        password=password,
    )


def load_rows(path: Path) -> list[UserRow]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv(path)
    if suffix == ".xlsx":
        return load_xlsx(path)
    raise KeycloakError(f"Unsupported file format: {path.suffix}. Use .csv or .xlsx")


def import_users(client: KeycloakAdminClient, rows: Iterable[UserRow], dry_run: bool) -> None:
    for user in rows:
        existing = None if dry_run else client.find_user_by_email(user.email)
        action = "create" if dry_run or not existing else "update"
        print(f"[plan] {action} {user.email}")

        if dry_run:
            continue

        if existing:
            user_id = existing["id"]
            client.update_user(user_id, user)
        else:
            user_id = client.create_user(user)

        client.set_password(user_id, user.password)
        print(f"[done] {action} {user.email} with UPDATE_PASSWORD")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import users into Keycloak from CSV/XLSX")
    parser.add_argument("file", type=Path, help="Path to the CSV or XLSX file")
    parser.add_argument("--base-url", default=os.getenv("KEYCLOAK_BASE_URL", "http://localhost:8003"))
    parser.add_argument("--realm", default=os.getenv("KEYCLOAK_REALM", "master"))
    parser.add_argument("--admin-user", default=os.getenv("KEYCLOAK_ADMIN_USER", "admin"))
    parser.add_argument("--admin-password", default=os.getenv("KEYCLOAK_ADMIN_PASSWORD"))
    parser.add_argument("--dry-run", action="store_true", help="Validate and print actions without calling Keycloak")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if not args.dry_run and not args.admin_password:
        parser.error("--admin-password is required unless --dry-run is used")

    try:
        rows = load_rows(args.file)
        print(f"Loaded {len(rows)} user(s) from {args.file}")

        client = None
        if not args.dry_run:
            client = KeycloakAdminClient(
                base_url=args.base_url,
                realm=args.realm,
                admin_user=args.admin_user,
                admin_password=args.admin_password,
            )

        import_users(client, rows, args.dry_run)
        print("Import completed successfully")
        return 0
    except KeycloakError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
